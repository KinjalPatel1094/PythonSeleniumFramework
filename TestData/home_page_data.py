import openpyxl


class HomePageData:
    test_HomePage_data = [{"firstname": "kinjal", "email": "test@gmail.com", "gender": "Female"},
                          {"firstname": "user", "email": "user@gmail.com", "gender": "Male"}]

    @staticmethod
    def get_test_data(test_case_name):
        book = openpyxl.load_workbook("/Users/kinjal/Documents/PythonExcelfile.xlsx")
        sheet = book.active

        Dict = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]

