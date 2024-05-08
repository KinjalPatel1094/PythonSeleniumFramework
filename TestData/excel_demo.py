import openpyxl

book = openpyxl.load_workbook("/Users/kinjal/Documents/PythonExcelfile.xlsx")
sheet = book.active  # sheet: <worksheet 'sheet1'>
cell = sheet.cell(row=1, column=2)
Dict = {}     # Dict: <class 'dict'>: {'firstname':'kinjal' ,'lastname':'patel'}
print(cell.value)
sheet.cell(row=2, column=2).value = "Kinjal"
print(sheet.cell(row=2, column=2).value)

print(sheet.max_row)
print(sheet.max_column)
print(sheet['A5'].value)

for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value == "testcase2":
        for j in range(2, sheet.max_column+1):

            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)



